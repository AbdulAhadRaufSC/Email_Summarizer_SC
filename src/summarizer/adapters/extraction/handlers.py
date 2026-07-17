"""Per-format text extraction handlers.

Each handler takes raw bytes and returns extracted text.  They raise on
failure — the caller (``SandboxedExtractor``) catches and degrades to
metadata-only.
"""

from __future__ import annotations

import csv
import io
import logging

logger = logging.getLogger(__name__)


def extract_pdf(data: bytes) -> str:
    """Extract text from a PDF using PyMuPDF (fitz)."""
    import fitz  # type: ignore[import-untyped]  # PyMuPDF

    text_parts: list[str] = []
    with fitz.open(stream=data, filetype="pdf") as doc:
        for page in doc:
            text_parts.append(page.get_text())
    return "\n".join(text_parts).strip()


def extract_docx(data: bytes) -> str:
    """Extract text from a DOCX file using python-docx."""
    from docx import Document

    doc = Document(io.BytesIO(data))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs).strip()


def extract_xlsx(data: bytes, *, max_rows: int = 50_000, max_cells: int = 500_000) -> str:
    """Extract text from an XLSX file using openpyxl.

    Enforces row and cell caps to prevent zip-bomb / memory exhaustion.
    """
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    total_cells = 0

    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"--- Sheet: {sheet} ---")
            for row_count, row in enumerate(ws.iter_rows(values_only=True)):
                if row_count >= max_rows:
                    lines.append(f"[Truncated: exceeded {max_rows} rows]")
                    break
                cells = [str(c) if c is not None else "" for c in row]
                total_cells += len(cells)
                if total_cells > max_cells:
                    lines.append(f"[Truncated: exceeded {max_cells} cells]")
                    break
                lines.append("\t".join(cells))
    finally:
        wb.close()

    return "\n".join(lines).strip()


def extract_csv_text(data: bytes) -> str:
    """Extract text from a CSV file using stdlib csv."""
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines: list[str] = []
    for i, row in enumerate(reader):
        if i >= 50_000:
            lines.append("[Truncated: exceeded 50000 rows]")
            break
        lines.append("\t".join(row))
    return "\n".join(lines).strip()


def extract_txt(data: bytes) -> str:
    """Extract text from a plain text file."""
    return data.decode("utf-8", errors="replace").strip()


def extract_image(data: bytes) -> str:
    """OCR an image using Tesseract via pytesseract.

    Requires the ``tesseract-ocr`` binary on the worker's PATH -- a
    system package, not something ``uv sync`` installs. Pillow's
    default ``Image.MAX_IMAGE_PIXELS`` guard raises
    ``DecompressionBombError`` on absurdly large images before OCR ever
    runs, which the caller's generic exception handling degrades to
    ``FAILED`` the same as any other extraction error.
    """
    import pytesseract  # type: ignore[import-untyped]
    from PIL import Image
    import os
    import sys

    # On Windows, tesseract is often not in PATH by default.
    if sys.platform == "win32":
        tesseract_paths = [
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Programs\Tesseract-OCR\tesseract.exe")
        ]
        for path in tesseract_paths:
            if os.path.exists(path):
                pytesseract.pytesseract.tesseract_cmd = path
                break

    with Image.open(io.BytesIO(data)) as image:
        text: str = pytesseract.image_to_string(image)
        return text.strip()
